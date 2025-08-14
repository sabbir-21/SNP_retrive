#@title **Retrieve RS then Mutation + Allele (one-click)**
#@markdown #<font color='orange'>**2**
import requests
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import re, time

# -----------------------
# Part A: Retrieve RS IDs (original code 1) with header row
# -----------------------
#@markdown Gene name Here eg: EGFR
filename = "tm6sf2" #@param {type:"string"}
#@markdown Choose filter type
missense = f"{filename}[All Fields] AND missense variant[Function_Class]"
missense_somatic = f"{filename}[All Fields] AND (missense variant[Function_Class] AND snp_snp_somatic[sb])"
filter = "missense_somatic" #@param ["missense", "missense_somatic"]
if filter == "missense":
    chosen_filter = missense
else:
    chosen_filter = missense_somatic
#@markdown Set maximum snp retrive
number = "2500" #@param {type:"string"}
url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
params = {
    "db": "snp",
    "term": f"{chosen_filter}",
    "retmax": f"{number}",
    "retmode": "json"
}

response = requests.get(url, params=params)
response.raise_for_status()
data = response.json()

id_list = data.get("esearchresult", {}).get("idlist", [])
rs_values = ["rs" + rs_id for rs_id in id_list]

workbook = Workbook()
sheet = workbook.active

# Add header row (row 1)
headers = [
    "RSID", "Molecule_type", "Change", "Codon", "Codon_Change",
    "Short_form", "Position", "Wild", "Mutant", "Chr_Position",
    "First", "Second", "Third"
]
sheet.append(headers)

for rs_id in rs_values:
    sheet.append([rs_id])

workbook.save(f"{filename}_missense.xlsx")
print(f"Done. Saved {len(rs_values)} rs IDs to {filename}_rs.xlsx")

# -----------------------
# Part B: Mutation + Allele Retrieve (original code 2)
# -----------------------
#@title **Mutation + Allele Retrieve**
# Open the existing Excel file
name = f'{filename}_missense.xlsx'
with open(name, 'rb') as tf:
    workbook = load_workbook(tf)
worksheet = workbook['Sheet']

# Iterate over column A in the 'Predict SNP' sheet
count = 0
for idx, cell in enumerate(worksheet['A'], start=1):
    # skip header row (row 1); modifications start from row 2
    if idx == 1:
        continue

    rsid = str(cell.value).strip()
    if rsid and rsid.lower() != 'none':
        count += 1
        print(count)
        url = f"https://www.ncbi.nlm.nih.gov/snp/{rsid}#variant_details"

        # pause to avoid hammering
        time.sleep(0.7)
        rr = requests.get(url)
        soup = BeautifulSoup(rr.text, 'html.parser')

        # -----------------------
        # Part 1: Missense Variant (original code1)
        # -----------------------
        table_rows = soup.find_all("tr")
        first_missense_change = None

        # Search for the first Missense Variant
        for row in table_rows:
            cells = row.find_all("td")

            if cells:
                # Check if the last cell contains "Missense Variant"
                if "Missense Variant" in cells[-1].get_text(strip=True):
                    # Extract the amino acid change from the third column (2nd <td>)
                    m_type = cells[0].get_text(strip=True)
                    first_missense_change_location = cells[1].get_text(strip=True)
                    first_missense_change = cells[2].get_text(strip=True)
                    break  # Stop after finding the first Missense Variant

        # If we found a Missense Variant, write it into column B-D and process E-G,H,I
        if first_missense_change:
            worksheet[f'B{idx}'] = m_type
            worksheet[f'C{idx}'] = first_missense_change_location
            worksheet[f'D{idx}'] = first_missense_change

            # Column E: extract last part after ':' (e.g., Met1Arg)
            match = re.search(r'p\.([A-Za-z]+\d+[A-Za-z]+)', first_missense_change_location)
            if match:
                aa_change = match.group(1)
                worksheet[f'E{idx}'] = aa_change

                # Column G: extract number from Met1Arg
                number_match = re.search(r'\d+', aa_change)
                if number_match:
                    worksheet[f'G{idx}'] = number_match.group()

                # Column F: build M1R from D and number
                d_col_value = worksheet[f'D{idx}'].value  # e.g., "M (Met) > R (Arg)"
                compact_match = re.match(r'([A-Z]) \(.*\) > ([A-Z]) \(.*\)', d_col_value)
                if compact_match and number_match:
                    from_aa = compact_match.group(1)
                    to_aa = compact_match.group(2)
                    num = number_match.group()
                    worksheet[f'F{idx}'] = f"{from_aa}{num}{to_aa}"

                    # Column H = first letter, Column I = last letter from F
                    worksheet[f'H{idx}'] = from_aa
                    worksheet[f'I{idx}'] = to_aa
        else:
            worksheet[f'B{idx}'] = 'No Missense Variant Found'

        # -----------------------
        # Part 2: Allele / Position (original code2)
        # -----------------------
        dl_tags = soup.find_all("dl")

        position = None
        alleles = None

        for dl in dl_tags:
            dts = dl.find_all("dt")
            dds = dl.find_all("dd")

            for dt, dd in zip(dts, dds):
                key = dt.text.strip()
                val = dd.text.strip()

                if key == "Position":
                    position = val.replace('\n', '').strip()
                elif key == "Alleles":
                    alleles = val.replace('\n', '').strip()

        if position and alleles:
            chr_match = re.search(r'chr(\d+|X|Y|MT):(\d+)', position)
            if chr_match:
                chrom = chr_match.group(1)
                pos = chr_match.group(2)

                # Split multiple alleles: A>C or C>A / C>T
                parts = re.split(r'\s*/\s*', alleles)
                formatted_alleles = []
                for allele in parts:
                    allele = allele.replace(">", ">").strip()
                    if ">" in allele:
                        ref, alt = allele.split(">")
                        formatted = f"{chrom},{pos},{ref},{alt}"
                        formatted_alleles.append(formatted)

                # Write position to column J
                worksheet[f'J{idx}'] = position

                # Write alleles to columns K,L,M (up to three)
                if len(formatted_alleles) > 0:
                    worksheet[f'K{idx}'] = formatted_alleles[0]
                if len(formatted_alleles) > 1:
                    worksheet[f'L{idx}'] = formatted_alleles[1]
                if len(formatted_alleles) > 2:
                    worksheet[f'M{idx}'] = formatted_alleles[2]

# Save once after processing all rows
workbook.save(name)
print(f"\n✅ Done! Processed {count} RSIDs and updated sheet '{worksheet.title}'.")