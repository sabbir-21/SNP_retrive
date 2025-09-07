import os
import re
import subprocess
import tempfile
import openpyxl

'''
Example excel: KEEP HEADER IN EXCEL
A        B
M1T      -1.886

'''

# Protein sequence
sequence = "MKVLWAALLVTFLAGCQAKVEQAVETEPEPELRQQTEWQSGQRWELALGRFWDYLRWVQTLSEQVQEELLSSQVTQELRALMDETMKELKAYKSELEEQLTPVAEETRARLSKELQAAQARLGADMEDVCGRLVQYRGEVQAMLGQSTEELRVRLASHLRKLRKRLLRDADDLQKRLAVYQAGAREGAERGLSAIRERLGPLVEQGRVRAATVGSLAGQPLQERAQAWGERLRARMEEMGSRTRDRLDEVKEQVAEVRAKLEEQAQQIRLQAEAFQARLKSWFEPLVEDMQRQWAGLVEKVQAAVGTSAAPVPSDNH"

# Path to MUpro installation and Excel file
mupro_dir = "."  # Current directory
excel_file = "ApoE_missense.xlsx"
short_form=6
saving_col=33
sheet_name="raw_filtered"
predict_script = os.path.join(mupro_dir, "bin", "predict_regr.sh")

# Function to parse mutation (e.g., 'M1R' → (1, 'M', 'R'))
def parse_mutation(mutation_str):
    match = re.match(r"([A-Z])(\d+)([A-Z])", mutation_str)
    if not match:
        raise ValueError(f"Invalid mutation format: {mutation_str}")
    orig_res, position, sub_res = match.groups()
    return int(position), orig_res, sub_res

# Function to create temporary input file for a mutation
def create_input_file(mutation, sequence):
    position, orig_res, sub_res = mutation
    name = f"{orig_res}{position}{sub_res}"
    input_content = f"{name}\n{sequence}\n{position}\n{orig_res}\n{sub_res}\n"
    with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".txt", dir=mupro_dir) as temp_file:
        temp_file.write(input_content)
        temp_file_path = temp_file.name
    return temp_file_path

# Function to run MUpro prediction and parse output
def run_prediction(input_file):
    try:
        result = subprocess.run(
            [predict_script, input_file],
            capture_output=True,
            text=True,
            check=True
        )
        output = result.stdout.strip()
        # Extract numerical delta delta G value
        match = re.search(r"Energy change \(delta G\) = ([\-\+]?[0-9]*\.?[0-9]+)", output)
        if match:
            return float(match.group(1))
        else:
            return f"Error: No delta G value found in output '{output}'"
    except subprocess.CalledProcessError as e:
        return f"Error: MUpro failed - {e.stderr}"
    finally:
        if os.path.exists(input_file):
            os.remove(input_file)

# Main function
def main():
    # Verify predict_regr.sh exists
    if not os.path.isfile(predict_script):
        print(f"Error: {predict_script} not found")
        return
    
    # Verify Excel file exists
    if not os.path.isfile(excel_file):
        print(f"Error: {excel_file} not found")
        return
    
    # Load Excel file
    try:
        wb = openpyxl.load_workbook(excel_file)
        if sheet_name not in wb.sheetnames:
            print("Error: Sheetname not found in Excel file")
            return
        sheet = wb[sheet_name]
    except Exception as e:
        print(f"Error loading Excel file: {str(e)}")
        return
    
    # Read mutations from column A and store results
    results = []
    b=0
    for row in sheet.iter_rows(min_row=2, min_col=short_form, max_col=short_form, values_only=True):
        b+=1
        print(b)
        mutation_str = row[0]
        if not mutation_str:
            continue
        try:
            position, orig_res, sub_res = parse_mutation(mutation_str)
            # Verify original residue
            if sequence[position - 1] != orig_res:
                results.append((mutation_str, f"Error: Expected {sequence[position - 1]} at position {position}"))
                continue
            # Create input file and run prediction
            input_file = create_input_file((position, orig_res, sub_res), sequence)
            ddg = run_prediction(input_file)
            results.append((mutation_str, ddg))
        except ValueError as e:
            results.append((mutation_str, f"Error: {str(e)}"))
    
    # Write results to column H (8th column)
    sheet.cell(row=1, column=saving_col).value = "Mupro"
    a=0
    for idx, (mutation_str, ddg) in enumerate(results, start=2):
        a+=1
        print(a)
        sheet.cell(row=idx, column=saving_col).value = ddg
    
    # Save Excel file
        wb.save(excel_file)
    
    

if __name__ == "__main__":
    main()