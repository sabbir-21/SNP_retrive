#@title **Auto DUET**

#!/usr/bin/env python3
"""
duet_batch_allthree.py
Bulk submit single-mutation requests to DUET and save mCSM, SDM and DUET results to Excel.

Requirements:
  pip install requests beautifulsoup4 openpyxl
"""

import requests
from bs4 import BeautifulSoup
import re
import time
import os
from openpyxl import Workbook, load_workbook

# -------- CONFIG ----------
MUTATION_LIST_FILE = "duet_1-220.txt"#@param {type:"string"}   # input list, one mutation per line
#@markdown duet.txt contains new line of M1R \n M1V
PDB_FILE = "model.pdb"#@param {type:"string"}            # pdb filename to upload
#@markdown pdb file of gene protein sequence. (SwissModel/Phyre)
CHAIN = "A" #@param {type:"string"}                      # hardcoded chain
#@markdown Chain of pdb dile by Discovery studio
OUT_XLSX = "duet_results_1-220.xlsx" #@param {type:"string"}   # output workbook
DUET_PAGE = "https://biosig.lab.uq.edu.au/duet/stability"
DUET_POST = "https://biosig.lab.uq.edu.au/duet/stability_prediction"
DELAY_BETWEEN = 2.0               # seconds between requests (politeness)
# --------------------------

def load_mutations(path):
    with open(path, 'r', encoding='utf-8') as f:
        muts = [line.strip() for line in f if line.strip()]
    return muts

def create_new_workbook(path):
    wb = Workbook()
    ws = wb.active
    # write headers in row 1
    headers = ["Mutation", "MCSM", "MCSM_class", "SDM", "SDM_class", "DUET", "DUET_class"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    wb.save(path)
    return wb, ws

def ensure_workbook(path):
    """Open existing workbook if it has the expected headers; otherwise create a new one."""
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            ws = wb.active
            # check header presence
            if (ws.cell(row=1, column=1).value and str(ws.cell(row=1, column=1).value).strip().lower() == "mutation"):
                return wb, ws
            else:
                # unexpected format -> create new workbook (overwrite)
                print("Existing workbook lacks expected headers — creating a new workbook and overwriting.")
                wb.close()
                return create_new_workbook(path)
        except Exception as e:
            print("Could not open existing workbook (will create new). Error:", e)
            return create_new_workbook(path)
    else:
        return create_new_workbook(path)

def save_wb(wb, path):
    wb.save(path)

def extract_number_and_label_from_font_tag(font_tag):
    """
    Given a <font> tag (BeautifulSoup), attempt to extract numeric ddg and the italic label.
    Returns (float_or_str_or_None, label_or_empty_str)
    """
    if font_tag is None:
        return None, ""
    # Try to extract numeric via regex
    text = font_tag.get_text(" ", strip=True)  # e.g. "-0.018 kcal/mol (Destabilizing)"
    m = re.search(r'([+-]?\d+(?:\.\d+)?)\s*kcal/mol', text)
    num = None
    if m:
        try:
            num = float(m.group(1))
        except Exception:
            num = m.group(1)
    # label: prefer <i> inside font
    i_tag = font_tag.find('i')
    label = ""
    if i_tag:
        label = i_tag.get_text(" ", strip=True)
    else:
        # fallback to parenthetical
        m2 = re.search(r'\(\s*([^)]+?)\s*\)', text)
        if m2:
            label = m2.group(1).strip()
    return num, label

def parse_three_predictions(html_text):
    """
    Parse the HTML and extract mCSM, SDM and DUET predicted ddG and labels.
    Returns dict: {'mcsm':(val,label), 'sdm':(val,label), 'duet':(val,label)}
    val is float if parseable else None; label is string (may be empty)
    """
    soup = BeautifulSoup(html_text, "html.parser")
    results = {'mcsm': (None, ""), 'sdm': (None, ""), 'duet': (None, "")}

    # iterate all <h4> tags and match by text
    for h4 in soup.find_all('h4'):
        txt = h4.get_text(" ", strip=True).lower()
        if 'mcsm predicted stability change' in txt:
            font = h4.find_next('font')
            results['mcsm'] = extract_number_and_label_from_font_tag(font)
        elif 'sdm predicted stability change' in txt:
            font = h4.find_next('font')
            results['sdm'] = extract_number_and_label_from_font_tag(font)
        elif 'duet' in txt and 'predicted stability change' in txt:
            # flatten text ensures we catch DUET even with nested <font>
            font = h4.find_next('font')
            # in some cases there may be multiple <font> tags immediately after h4
            # pick the first one that contains 'kcal/mol'
            fonts = h4.find_all_next('font', limit=3)
            for f in fonts:
                if 'kcal/mol' in f.get_text():
                    font = f
                    break
            results['duet'] = extract_number_and_label_from_font_tag(font)

    # final safety: if any still None, attempt regex global search (fallback)
    # pattern: number ... (Label)
    if results['mcsm'][0] is None:
        # try to locate block that contains "mCSM Predicted"
        block = soup.find_all(string=re.compile(r"mCSM Predicted Stability Change", re.I))
        if block:
            parent = block[0].parent
            font = parent.find_next('font')
            results['mcsm'] = extract_number_and_label_from_font_tag(font)
    if results['sdm'][0] is None:
        block = soup.find_all(string=re.compile(r"SDM Predicted Stability Change", re.I))
        if block:
            parent = block[0].parent
            font = parent.find_next('font')
            results['sdm'] = extract_number_and_label_from_font_tag(font)
    if results['duet'][0] is None:
        block = soup.find_all(string=re.compile(r"DUET Predicted Stability Change", re.I))
        if block:
            parent = block[0].parent
            font = parent.find_next('font')
            results['duet'] = extract_number_and_label_from_font_tag(font)

    return results

def main():
    # load mutation list
    if not os.path.exists(MUTATION_LIST_FILE):
        print(f"Mutation list file not found: {MUTATION_LIST_FILE}")
        return
    mutations = load_mutations(MUTATION_LIST_FILE)
    if not mutations:
        print("No mutations found in the input file.")
        return

    # prepare workbook (create or validate headers)
    wb, ws = ensure_workbook(OUT_XLSX)

    session = requests.Session()
    # initial GET to fetch cookies / possible CSRF
    try:
        session.get(DUET_PAGE, timeout=20)
    except Exception as e:
        print("Warning: initial GET to DUET page failed:", e)

    # Start writing at row 2 (row 1 has headers)
    start_row = 2

    for idx, mut in enumerate(mutations, start=start_row):
        print(f"[{idx - start_row + 1}/{len(mutations)}] Processing {mut} -> Excel row {idx}")
        # defaults in case of error
        mcsm_val = "error"
        mcsm_label = ""
        sdm_val = "error"
        sdm_label = ""
        duet_val = "error"
        duet_label = ""

        try:
            if not os.path.exists(PDB_FILE):
                raise FileNotFoundError(f"PDB file not found: {PDB_FILE}")

            with open(PDB_FILE, 'rb') as pdb_f:
                files = {
                    'wild': (os.path.basename(PDB_FILE), pdb_f, 'application/octet-stream'),
                }
                data = {
                    'pdb_code': '',
                    'mutation': mut,
                    'chain': CHAIN,
                    'run': 'single',
                    'mutation_sys': '',
                    'chain_sys': '',
                }
                headers = {
                    'Referer': DUET_PAGE,
                    'User-Agent': 'python-requests/duet-batch-script',
                }
                resp = session.post(DUET_POST, data=data, files=files, headers=headers, timeout=120)
                resp.raise_for_status()
                html = resp.text

                parsed = parse_three_predictions(html)

                # assign parsed values if present, else 'error'
                mcsm_v, mcsm_l = parsed.get('mcsm', (None, ""))
                sdm_v, sdm_l = parsed.get('sdm', (None, ""))
                duet_v, duet_l = parsed.get('duet', (None, ""))

                if mcsm_v is not None:
                    mcsm_val = mcsm_v
                    mcsm_label = mcsm_l
                else:
                    mcsm_val = "error"
                    mcsm_label = mcsm_l or ""

                if sdm_v is not None:
                    sdm_val = sdm_v
                    sdm_label = sdm_l
                else:
                    sdm_val = "error"
                    sdm_label = sdm_l or ""

                if duet_v is not None:
                    duet_val = duet_v
                    duet_label = duet_l
                else:
                    duet_val = "error"
                    duet_label = duet_l or ""

        except Exception as ex:
            print(f"Error processing {mut}: {ex}")
            mcsm_val = sdm_val = duet_val = "error"
            mcsm_label = sdm_label = duet_label = ""

        # Write to excel ensuring exact row mapping
        try:
            ws.cell(row=idx, column=1, value=mut)         # A
            ws.cell(row=idx, column=2, value=mcsm_val)    # B
            ws.cell(row=idx, column=3, value=mcsm_label)  # C
            ws.cell(row=idx, column=4, value=sdm_val)     # D
            ws.cell(row=idx, column=5, value=sdm_label)   # E
            ws.cell(row=idx, column=6, value=duet_val)    # F
            ws.cell(row=idx, column=7, value=duet_label)  # G
            save_wb(wb, OUT_XLSX)
        except Exception as ex_save:
            print(f"Failed to write/save Excel for row {idx}: {ex_save}")

        time.sleep(DELAY_BETWEEN)

    print("Done. Results saved to", OUT_XLSX)

if __name__ == "__main__":
    main()