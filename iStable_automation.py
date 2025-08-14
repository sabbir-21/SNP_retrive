#@title **Auto istable**
"""
iStable batch submit + result writer

- Input Excel (Sheet1) must have headers: Position, Wild, Mutant
- Outputs written back into same sheet in columns:
    D -> Conf_Score
    E -> iStable

Set INPUT_FILE and SEQ below, then run.
"""
import requests
import pandas as pd
import time
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from uuid import uuid4

# ---------- USER SETTINGS ----------
INPUT_FILE = "istable_calculator.xlsx"#@param {type:"string"}   # path to your Excel file
#@markdown contains new lines of 'Position'	'Wild'	'Mutant'
#@markdown ___
'''
Position	Wild	Mutant
1	M	R
'''
SHEET_NAME = "Sheet1"#@param {type:"string"}
SEQ = ("MDIPPLAGKIAALSLSALPVSYALNHVSALSHPLWVALMSALILGLLFVAVYSLSHGEVSYDPLYAVFAVFAFTSVVDLIIALQEDSY"
       "VVGFMEFYTKEGEPYLRTAHGVFICYWDGTVHYLLYLAMAGAICRRKRYRNFGLYWLGSFAMSILVFLTGNILGKYSSEIRPAFFLTIPY"
       "LLVPCWAGMKVFSQPRALTRCTANMVQEEQRKGLLQRPADLALVIYLILAGFFTLFRGLVVLDCPTDACFVYIYQYEPYLRDPVAYPKVQ"
       "MLMYMFYVLPFCGLAAYALTFPGCSWLPDWALVFAGGIGQAQFSHMGASMHLRTPFTYRVPEDTWGCFFVCNLLYALGPHLLAYRCLQWPA"
       "FFHQPPPSDPLALHKKQH")
TEMP = "25"#@param {type:"string"}
PH = "7"#@param {type:"string"}
URL = "http://predictor.nchu.edu.tw/iStable/indexSeq.php"
# -----------------------------------

HEADERS = {
    "User-Agent": "python-requests/2.x",
    "Referer": "http://predictor.nchu.edu.tw/iStable/indexSeq.php",
    "Content-Type": "application/x-www-form-urlencoded"
}

def parse_istable_response(html_text):
    """
    Parse the returned HTML and extract the iStable fusion result and confidence score.
    Returns (result_text_or_None, conf_score_or_None)
    """
    soup = BeautifulSoup(html_text, "lxml")

    # Strategy: find the table with the 'Predictor' header, then the row whose first cell is 'Result'
    for table in soup.find_all("table"):
        # require a table that mentions Predictor somewhere
        if table.find(text=re.compile(r'Predictor', re.I)):
            # find the row labelled 'Result'
            for tr in table.find_all("tr"):
                first_td = tr.find("td")
                if first_td and first_td.get_text(strip=True).lower() == "result":
                    tds = tr.find_all("td")
                    # Expect at least 7 tds like in the sample HTML
                    if len(tds) >= 7:
                        istable_val = tds[5].get_text(strip=True)
                        conf_val = tds[6].get_text(strip=True)
                        # Normalize empty/nulls
                        if istable_val.lower() in ("", "null", "na", "n/a"):
                            istable_val = None
                        if conf_val.lower() in ("", "null", "na", "n/a"):
                            conf_val = None
                        return istable_val, conf_val
    # Fallbacks: try to search for "Fusion Result" or "iStable" text anywhere
    # (less reliable)
    txt = soup.get_text(" ", strip=True)
    m = re.search(r'(Fusion Result[:\s]*)(Increase|Decrease|Neutral)', txt, re.I)
    conf = None
    if m:
        return m.group(2), None
    m2 = re.search(r'([0-9]+\.[0-9]+)', txt)
    if m2:
        conf = m2.group(1)
    return None, conf

def build_payload(jobname, wild, position, mutant, seq, temp=TEMP, ph=PH):
    """
    Build the form payload for indexSeq.php based on observed fields.
    """
    return {
        "jobname": jobname,
        "wildtype": wild,
        "position": str(position),
        "mutant": mutant,
        "temp": str(temp),
        "ph": str(ph),
        "seq": seq
    }

def main():
    # read excel
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, dtype=str)
    # normalize header names (strip)
    df.columns = [c.strip() for c in df.columns]

    # ensure required columns exist
    for col in ("Position", "Wild", "Mutant"):
        if col not in df.columns:
            raise ValueError(f"Input Excel must have column '{col}'")

    # open workbook for writing results
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    session = requests.Session()
    session.headers.update(HEADERS)

    # We'll assume header row is row 1. Find column indexes for existing headers.
    header_row = 1
    col_index = {}
    for col in ws[1]:
        header = (col.value or "").strip()
        col_index[header] = col.column  # openpyxl column (int or letter depending on version)
    # Determine where to write results: D (Conf_Score) and E (iStable) per user's request.
    # If those headers don't exist, write them into D1 and E1 (overwriting if present).
    def ensure_header(name, col_letter):
        if name not in col_index:
            ws[f"{col_letter}1"] = name
            # update mapping
            col_index[name] = ws[f"{col_letter}1"].column

    # Use columns D and E explicitly
    ensure_header("Conf_Score", "D")
    ensure_header("iStable", "E")

    # iterate rows in dataframe (pandas keeps original order)
    for idx, row in df.iterrows():
        # Excel row number = idx + 2 (because pandas index starts at 0 and header occupies row 1)
        excel_row = idx + 2
        pos = row["Position"]
        wild = row["Wild"]
        mut = row["Mutant"]

        # Skip if already has value (optional). Here we overwrite anyway.
        jobname = f"job_{uuid4().hex[:8]}_{wild}{pos}{mut}"
        payload = build_payload(jobname, wild, pos, mut, SEQ, TEMP, PH)

        try:
            resp = session.post(URL, data=payload, timeout=30)
            if resp.status_code != 200:
                print(f"Row {excel_row}: HTTP {resp.status_code} for {wild}{pos}{mut}")
                istable_val, conf_val = None, None
            else:
                istable_val, conf_val = parse_istable_response(resp.text)
                print(f"Row {excel_row}: parsed -> iStable={istable_val!r}, Conf={conf_val!r}")
        except Exception as e:
            print(f"Row {excel_row}: request error: {e}")
            istable_val, conf_val = None, None

        # write back to sheet columns "Conf_Score" and "iStable"
        conf_col = col_index["Conf_Score"]
        ist_col = col_index["iStable"]

        # openpyxl expects numeric column indexes; ensure we use column letter or number properly
        # ws.cell(row, column) uses numeric column index
        # col_index returns numeric column if using openpyxl's .column attribute
        if isinstance(conf_col, str):
            # convert letter to number
            from openpyxl.utils import column_index_from_string
            conf_col_num = column_index_from_string(conf_col)
        else:
            conf_col_num = conf_col

        if isinstance(ist_col, str):
            from openpyxl.utils import column_index_from_string
            ist_col_num = column_index_from_string(ist_col)
        else:
            ist_col_num = ist_col

        # Write values (strings). For missing values we write empty string.
        ws.cell(row=excel_row, column=conf_col_num, value=(conf_val or ""))
        ws.cell(row=excel_row, column=ist_col_num, value=(istable_val or ""))

        # be polite to server — short pause
        time.sleep(0.8)

    
    wb.save(INPUT_FILE)
    print(f"Results written to {INPUT_FILE}")

if __name__ == "__main__":
    main()
