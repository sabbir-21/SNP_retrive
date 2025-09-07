'''
Example excel:

D           E       F
Position    New     DDG
12          V       0.87
'''

##############################
import openpyxl
import subprocess
import sys

# Python 2.7 compatible script
# Assuming openpyxl is installed: pip install openpyxl==2.5.14 for Python 2.7 support

# Excel file name - change if needed
excel_file = 'ApoE_missense.xlsx'  # Replace with your actual file name
position=7 # Column Position
mutant=9 # Column Mutant
saving_col=32 # Column Saving
# Load the workbook
wb = openpyxl.load_workbook(excel_file)
sheet = wb['raw_filtered']

# Find the total number of rows with data (assuming header in row 1, data from row 2)
max_row = sheet.max_row
total_mutations = max_row - 0  # Assuming row 1 is header

# Counters
done = 0

# Loop through rows starting from 2
for row in range(1, max_row + 1):
    pos_cell = sheet.cell(row=row, column=position)  # Column Position (4)
    new_res_cell = sheet.cell(row=row, column=mutant)  # Column Mutant (5)
    
    pos = pos_cell.value
    new_res = new_res_cell.value
    
    # Skip if position or new_res is empty
    if pos is None or new_res is None:
        continue
    
    # Convert pos to string for command
    pos_str = str(pos)
    
    # Build the command
    cmd = [
        'python', '-O', 'I-Mutant2.0.py',
        '-seqv', 'a.seq',
        pos_str, new_res,
        'pH=7.0', 'Temperature=25.0'
    ]
    
    try:
        # Run the command and capture output
        output = subprocess.check_output(cmd)
        
        # Parse the output to find DDG
        ddg = None
        for line in output.splitlines():
            stripped = line.strip()
            if stripped.startswith(pos_str):
                parts = stripped.split()
                if len(parts) >= 4:
                    ddg = parts[3]
                    break
        
        if ddg is not None:
            # Write to column F (6)
            sheet.cell(row=row, column=saving_col).value = float(ddg)
            
            # Increment done
            done += 1
            
            # Print progress
            remaining = total_mutations - done
            print "Processed mutation at row {}: Done {}, Remaining {}".format(row, done, remaining)
            
            # Save the workbook after each mutation
            wb.save(excel_file)
        else:
            print "Warning: DDG not found in output for row {}".format(row)
    
    except subprocess.CalledProcessError as e:
        print "Error running command for row {}: {}".format(row, e)
    except ValueError:
        print "Error parsing DDG as float for row {}".format(row)

# Final save (redundant but safe)
wb.save(excel_file)
print "All processing complete. Total done: {}".format(done)